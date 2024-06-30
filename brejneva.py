import pandas as pd
import xlsxwriter
import openpyxl
import re
from collections import OrderedDict


def remove_leading_zeros(s):
    return re.sub(r'^0+', '', s)

# Путь к вашему файлу Excel
file_path = 'kniit.xlsx'

# Чтение данных из файла Excel
df = pd.read_excel(file_path)

# Выбор диапазона ячеек от A до AH
selected_cells = df.iloc[:, :35]  # Предполагается, что у вас есть 10 столбцов от A до J
selected_cells = selected_cells[2:]
# Преобразование выбранного диапазона в список
cell_list = selected_cells.values.tolist()

# print(cell_list)
# print(cell_list)


class Abitura:
    def __init__(self, row=[], prior=[], tables=[], dop=[]):
        self.row = row
        self.prior = prior
        self.tables = tables
        self.dop = dop


kniit = {
        "Информатика и вычислительная техника": "ИВТ",
        "Математическое обеспечение и администрирование информационных систем": "МОАИС",
        "Программная инженерия": "ПИ",
        "Компьютерная безопасность": "КБ",
        "Фундаментальная информатика и информационные технологии": "ФИИТ",
        "Системный анализ и управление": "САУ"

    }

workbook = xlsxwriter.Workbook('kniit2.xlsx')
worksheet = workbook.add_worksheet()



abituras_old = {}
abituras_old["budj_obs"] = OrderedDict()
abituras_old["budj_osob"] = OrderedDict()
abituras_old["cel"] = OrderedDict()
abituras_old["kom"] = OrderedDict()
abituras_old["zo"] = OrderedDict()
abituras_old["inostr"] = OrderedDict()


file_path = 'kniit.xlsx'

for fl in abituras_old:
    wb = openpyxl.load_workbook("tables/" + fl + ".xlsx")
    sheet = wb.active
    for i in range(1, sheet.max_row + 1):
        row_data = []
        for j in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=i, column=j).value
            if cell_value is not None:
                row_data.append(cell_value)
            else:
                row_data.append('')
        abituras_old[fl][row_data[0]] = row_data




h = 0
abituras_arr=[]
abituras = {}

for row in cell_list:
    row12 = row[12].split('_')[2:]
    prior = ['']*6
    pr = 0
    tables = set()
    flag1 = True
    if row12[-1][:3] in ["Бак", "Спе"]:
        if row[10] > 5:
            # prior[5].append(kniit[row[8]])
            pr = 5
        else:
            pr = row[10]-1
            # prior[pr] = kniit[row[8]]
        prior[pr] += kniit[row[8]]+" "
        # print(prior)
        print(row12)
        if (row12[0]) == "З":
            prior[pr] += "ЗО "
            tables.add("zo")
            flag1 = False
        if "и" in row12:
            prior[pr] += "ИК "
            tables.add("inostr")
        if "ОК" in row12:
            prior[pr] += "ОК "
            tables.add("budj_osob")
        elif "ОП" in row12:
            prior[pr] += "ОП "
            tables.add("budj_osob")
        elif "ПО" in row12:
            prior[pr] += "К "
            tables.add("kom")
        elif "Ц" in row12:
            prior[pr] += "Ц "
            tables.add("cel")
        elif flag1:
            tables.add("budj_obs")
        # print(prior)
        # print(tables)

        if not isinstance(row[6],str):
            row[6] = remove_leading_zeros(str(row[5]))
        zam =''
        lich = "лично"
        orig = "Копия"
        if row[21] == "Нет документа об образовании":
            zam = row[21]
        if row[31] != "Лично":
            lich = "госуслуги"
        if row[32] == "Оригинал":
            orig = row[32]
        row_ab = [row[4],row[6],'',orig,'','',lich, zam]
        # print(row_ab, row[10])
        if row_ab[1] != "161-452-197 48":
            if row[6] not in abituras:
                row_ab[0], row_ab[1] = row_ab[1], row_ab[0]
                abituras[row[6]] = Abitura(row=row_ab, prior=prior, tables=tables)
                flag = False
                for fl in abituras_old:

                    if row[6] in abituras_old[fl]:
                        if abituras_old[fl][row[6]][1] != '':
                            flag = True
                if not flag:
                    worksheet.write(h, 1, row[4])
                    worksheet.write(h, 4, row[33])
                    h += 1
            else:
                for i in range(6):
                    abituras[row[6]].prior[i] += prior[i]

workbook.close()

workbook = xlsxwriter.Workbook('new.xlsx')
worksheet = workbook.add_worksheet()


workbook = {}
worksheet = {}


for fl in abituras_old:
    i = 0
    workbook = xlsxwriter.Workbook("tables/"+fl+'.xlsx')
    worksheet = workbook.add_worksheet()
    for ab in abituras_old[fl]:
        if len(abituras_old[fl][ab]) < 20:
            abituras_old[fl][ab] = abituras_old[fl][ab] + [''] * (20 - len(abituras_old[fl][ab]))
        abitr = abituras_old[fl][ab]
        if ab not in abituras:
            for j in range(len(abitr)):
                worksheet.write(i, j, abitr[j])
        else:
            if abituras[ab].row[-1] == "":
                abituras_old[fl][ab][-1].replace('Нет документа об образовании', "")
            row = abituras_old[fl][ab][:3] + abituras[ab].prior  +abituras_old[fl][ab][9:]
            for j in range(len(row)):
                worksheet.write(i, j, row[j])
        i+=1

    for ab in abituras:
        if fl in abituras[ab].tables and ab not in abituras_old[fl]:
            row = abituras[ab].row[:3] + abituras[ab].prior + [""] * 6 + abituras[ab].row[3:]
            # print(row)
            for j in range(len(row)):
                # print( row[j])
                worksheet.write(i, j, row[j])
            i +=1
    workbook.close()
#
#
# for ab in abituras:
#     print(abituras[ab].row, abituras[ab].prior)

 # workbook.close()
