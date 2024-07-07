import requests
from bs4 import BeautifulSoup
import re
import openpyxl
from collections import OrderedDict
import xlsxwriter


def remove_leading_zeros(s):
    return re.sub(r'^0+', '', s)


urls = ["https://www.sgu.ru/svodka/1295", "https://www.sgu.ru/svodka/1494", "https://www.sgu.ru/svodka/1619",

        "https://www.sgu.ru/svodka/1353", "https://www.sgu.ru/svodka/2404", "https://www.sgu.ru/svodka/1294",
        "https://www.sgu.ru/svodka/1495", "https://www.sgu.ru/svodka/1620",
        "https://www.sgu.ru/svodka/1354", "https://www.sgu.ru/svodka/2396", "https://www.sgu.ru/svodka/2397",
        "https://www.sgu.ru/svodka/1293", "https://www.sgu.ru/svodka/1496", "https://www.sgu.ru/svodka/1621",

        "https://www.sgu.ru/svodka/1355", "https://www.sgu.ru/svodka/2402", "https://www.sgu.ru/svodka/1292",
        "https://www.sgu.ru/svodka/1497", "https://www.sgu.ru/svodka/1622",
        "https://www.sgu.ru/svodka/1356", "https://www.sgu.ru/svodka/2398",
        "https://www.sgu.ru/svodka/2399", "https://www.sgu.ru/svodka/1436", "https://www.sgu.ru/svodka/1500",
        "https://www.sgu.ru/svodka/1625", "https://www.sgu.ru/svodka/1437",
        "https://www.sgu.ru/svodka/1291", "https://www.sgu.ru/svodka/1624",
        "https://www.sgu.ru/svodka/1358", "https://www.sgu.ru/svodka/2413", "https://www.sgu.ru/svodka/2414",
        "https://www.sgu.ru/svodka/2415",
        "https://www.sgu.ru/svodka/1296", "https://www.sgu.ru/svodka/1498", "https://www.sgu.ru/svodka/1623",

        "https://www.sgu.ru/svodka/1357", "https://www.sgu.ru/svodka/1499", "https://www.sgu.ru/svodka/1787", "https://www.sgu.ru/svodka/1788", "https://www.sgu.ru/svodka/1791",
        "https://www.sgu.ru/svodka/1792", "https://www.sgu.ru/svodka/1795", "https://www.sgu.ru/svodka/1794",
        "https://www.sgu.ru/svodka/1793"
        ]

abituras = {}
abituras["budj_obs"] = OrderedDict()
abituras["budj_osob"] = OrderedDict()
abituras["cel"] = OrderedDict()
abituras["kom"] = OrderedDict()
abituras["zo"] = OrderedDict()

abituras_old = {}
abituras_old["budj_obs"] = OrderedDict()
abituras_old["budj_osob"] = OrderedDict()
abituras_old["cel"] = OrderedDict()
abituras_old["kom"] = OrderedDict()
abituras_old["zo"] = OrderedDict()

st_accept = "text/html"
st_useragent = "Mozilla/5.0 (Macintosh; Intel Mac OS X 12_3_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.4 Safari/605.1.15"
pattern = r"\d{1,3}\.\d{1,3}\.\d{1,3}"
pattern2 = r'(Бюджет/Общие|Бюджет/Особое|Бюджет/Отдельная|Полное|Целевой)'
kniit = {
    "Информатика и вычислительная техника": "ИВТ",
    "Математическое обеспечение и администрирование информационных систем": "МОАИС",
    "Программная инженерия": "ПИ",
    "Компьютерная безопасность": "КБ",
    "Фундаментальная информатика и информационные технологии": "ФИИТ",
    "Системный анализ и управление": "САУ"

}
not_kniit = {
    "Бизнес-информатика": "БИ",
    "Информационные системы и технологии": "ИСТ",
    "Прикладная математика и информатика": "ПМИ",
    "Математика и компьютерные науки": "МКН",
    "Прикладная информатика социологии": "ПРИ соц",
    "Прикладная информатика экономике": "ПРИ экон мех",
    "Прикладная информатика": "ПРИ мех",
    "Конструирование и технология электронных средств": "КТЭС",
    "Биотехнические системы и технологии": "БСТ",
    "Техносферная безопасность": "ТБ",
    "Геология": "Геол",
    "Механика и математическое моделирование": "МиММ",
    "Электроника и наноэлектроника": "ЭиН",
    "Экономика": "Экон",
    "Таможенное дело": "Тамож",
    "Менеджмент": "Менедж",
    "Медицинская биофизика": "МБ",
    "Медицинская кибернетика": "МК",
    "Прикладная геология": "ПГ",
    "Инфокоммуникационные технологии и системы связи": "ИТиСС",
    "Педагогическое образование Информатика": "ПО инф",
    "Педагогическое образование Математическое": "ПО мат",
    "Психолого-педагогическое образование": "ППО",
    "Социология": "Соц",
    "Сервис": "Серв"
}

for url in urls:
    print(url)
    req = requests.get(url)
    src = req.text
    soup = BeautifulSoup(src, 'lxml')
    tables = soup.find_all('table')

    for table in tables:
        for row in table.find_all('tr'):
            cols = row.find_all('td')
            cols = [col.text.strip() for col in cols]
            fin_row = [''] * 20
            if len(cols) > 9:
                cols[1] = remove_leading_zeros(cols[1])
                ttl = soup.find_all(class_="svodka-table__info-wrap")[0]
                ttl = ttl.find_all(class_="svodka-table__info-name")
                ttl = (ttl[2].string, ttl[4].string, ttl[5].string.split(' ')[0])

                name_first = kniit[ttl[0]]
                osnov = "ОО"
                fl = "budj_obs"

                if ttl[1] == "Заочная":
                    name_first += " ЗО"
                    if ttl[2] != "Полное":
                        fl = "zo"
                if ttl[2] == "Бюджет/Особое":
                    fl = "budj_osob"
                    osnov = "ОП"
                if ttl[2] == "Бюджет/Отдельная":
                    fl = "budj_osob"
                    osnov = "ОК"
                if ttl[2] == "Полное":
                    fl = "kom"
                    osnov = "К"

                if ttl[2] == "Целевой":
                    fl = "cel"
                    osnov = "Ц"
                cols[9] = int(cols[9])
                if cols[1] in abituras[fl]:
                    abituras[fl][cols[1]][cols[9] + 2] = name_first
                else:
                    cols[10] = ' '.join(cols[10].split())[4:]

                    i = 0
                    s = cols[10].split()

                    if cols[9] > 6:
                        cols[9] = 6
                    fin_row16 = ""
                    while i < len(s):
                        if re.match(pattern, s[i]) is not None:
                            pr = int(s[i:][int(s[i:].index('Приоритет:') + 1)]) - 1
                            name = ""
                            j = i + 1
                            flag = True
                            while re.match(pattern2, s[j]) is None:
                                if s[j] == "(Профиль:":
                                    flag = False
                                    if "Прикладная информатика" in name:
                                        name += " " + s[j + 4]
                                    if "Педагогическое образование" in name:
                                        name += " " + s[j + 1]

                                if flag:
                                    name += " " + s[j]
                                j += 1
                            name = name[1:]
                            if name in not_kniit:
                                fin_row16 += str(pr + 1) + " - " + not_kniit[name]
                                if "(заоч.)" == s[j + 2]:
                                    fin_row16 += " ЗО"
                                elif s[j + 2] == "затрат" and "(заоч.)" == s[j + 3]:
                                    fin_row16 += " ЗО"
                                if s[j] == "Бюджет/Особое":
                                    fin_row16 += " ОП"
                                elif s[j] == "Бюджет/Отдельная":
                                    fin_row16 += " ОК"
                                elif s[j] == "Полное":
                                    fin_row16 += " К"
                                elif s[j] == "Целевой":
                                    fin_row16 += " Ц"
                                fin_row16 += " "
                            i = j
                        i += 1
                    orig = "Копия"
                    if "Ориг." in s:
                        orig = "Оригинал"
                    fin_row = [cols[1], "", osnov] + [""] * 6 + cols[2:7]+ [""] + [orig, fin_row16, '', "госуслуги", "Нет документа об образовании"]
                    fin_row = ['0' if item in ['Нет данных', '---'] else item for item in fin_row]
                    fin_row[cols[9] + 2] = name_first
                    abituras[fl][cols[1]] = fin_row

for fl in abituras_old:
        wb = openpyxl.load_workbook("tables/" + fl + ".xlsx")
        sheet = wb.active
        for i in range(1, sheet.max_row + 1):
            row_data = []
            for j in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=i, column=j).value
                if cell_value is not None:
                    row_data.append(cell_value)
                else:
                    row_data.append('')
            abituras_old[fl][remove_leading_zeros(row_data[0])] = row_data


for fl in abituras:
        g = 0
        workbook = xlsxwriter.Workbook("tables/" + fl + '.xlsx')
        worksheet = workbook.add_worksheet()
        for ab in abituras_old[fl]:
            if ab in abituras:
                if abituras_old[fl][ab][15] != abituras[fl][ab][15]:
                    print(ab, ": ", abituras_old[fl][ab][15], " -> ", abituras[ab][fl][ab][15])
                    if abituras_old[fl][ab][13] != '':
                        if int(abituras_old[fl][ab][13]) + int(abituras_old[fl][ab][14]) < int(
                                abituras[fl][ab][13]) + int(abituras[fl][ab][14]):
                            print(ab, int(abituras_old[fl][ab][13]) + int(abituras_old[fl][ab][14]) - int(
                                abituras[fl][ab][13]) + int(abituras[fl][ab][14]))
                    if abituras[ab].row[7] == "Оригинал":
                        for num in range(5):
                            if abituras_old[fl][ab][3 + num][:2] != abituras[fl][ab][3 + num][:2]:
                                print(ab, abituras_old[fl][ab][3 + num], "-> ", abituras[ab][fl][ab][3 + num][:2])

                    abituras_old[fl][ab][2] = abituras[fl][ab][2]
                    for num in range(6):
                        abituras_old[fl][ab][2 + num] = abituras[fl][ab][2 + num]
                    abituras_old[fl][ab][16] = abituras[fl][ab][16]
            abitr = abituras_old[fl][ab]
            for j in range(len(abitr)):
                worksheet.write(g, j, abitr[j])
            g += 1
        for ab in abituras[fl]:
            if ab not in abituras_old:
                print(abituras[fl][ab])
                abitr = abituras[fl][ab]
                for j in range(len(abitr)):
                    worksheet.write(g, j, abitr[j])
            g += 1
        workbook.close()
