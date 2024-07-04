import requests
from bs4 import BeautifulSoup
import re
import openpyxl
from collections import OrderedDict
import xlsxwriter

def remove_leading_zeros(s):
    return re.sub(r'^0+', '', s)



urls = ["https://www.sgu.ru/svodka/1295", "https://www.sgu.ru/svodka/1494", "https://www.sgu.ru/svodka/1619",

        "https://www.sgu.ru/svodka/1353", "https://www.sgu.ru/svodka/2404", "https://www.sgu.ru/svodka/1294",
        "https://www.sgu.ru/svodka/1495", "https://www.sgu.ru/svodka/1620",
        "https://www.sgu.ru/svodka/1354", "https://www.sgu.ru/svodka/2396", "https://www.sgu.ru/svodka/2397",
        "https://www.sgu.ru/svodka/1293", "https://www.sgu.ru/svodka/1496", "https://www.sgu.ru/svodka/1621",

        "https://www.sgu.ru/svodka/1355", "https://www.sgu.ru/svodka/2402", "https://www.sgu.ru/svodka/1292",
        "https://www.sgu.ru/svodka/1497", "https://www.sgu.ru/svodka/1622",
        "https://www.sgu.ru/svodka/1356", "https://www.sgu.ru/svodka/2398",
        "https://www.sgu.ru/svodka/2399", "https://www.sgu.ru/svodka/1436", "https://www.sgu.ru/svodka/1500",
        "https://www.sgu.ru/svodka/1625",  "https://www.sgu.ru/svodka/1437",
        "https://www.sgu.ru/svodka/1291", "https://www.sgu.ru/svodka/1624",
        "https://www.sgu.ru/svodka/1358", "https://www.sgu.ru/svodka/2413", "https://www.sgu.ru/svodka/2414",
        "https://www.sgu.ru/svodka/2415",
        "https://www.sgu.ru/svodka/1296", "https://www.sgu.ru/svodka/1498", "https://www.sgu.ru/svodka/1623",

        "https://www.sgu.ru/svodka/1357", "https://www.sgu.ru/svodka/1499"
        ]


class Abitura:
    def __init__(self, row, prior, tables, dop):
        self.row = row
        self.prior = prior
        self.tables = tables
        self.dop = dop


def func(spisok, urls):
    abituras = {}

    abituras_old = {}
    abituras_old["budj_obs"] = OrderedDict()
    abituras_old["budj_osob"] = OrderedDict()
    abituras_old["cel"] = OrderedDict()
    abituras_old["kom"] = OrderedDict()
    abituras_old["zo"] = OrderedDict()

    d = {}
    d["budj_obs"] = ["Раздел ОБЩИЙ КОНКУРС"]
    d["budj_osob"] =  ["Раздел Особые Права (квота и спецквота)"]
    d["cel"] = ["Раздел Целевики"]
    d["kom"] = ["Коммерция"]
    d["zo"] = ["ПИ ЗАОЧНОЕ"]


    st_accept = "text/html"
    st_useragent = "Mozilla/5.0 (Macintosh; Intel Mac OS X 12_3_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.4 Safari/605.1.15"
    headers = {
        "Accept": st_accept,
        "User-Agent": st_useragent
    }

    pattern = r"\d{1,3}\.\d{1,3}\.\d{1,3}"
    pattern2 = r'(Бюджет/Общие|Бюджет/Особое|Бюджет/Отдельная|Полное|Целевой)'
    kniit = {
        "Информатика и вычислительная техника": "ИВТ",
        "Математическое обеспечение и администрирование информационных систем": "МОАИС",
        "Программная инженерия": "ПИ",
        "Компьютерная безопасность": "КБ",
        "Фундаментальная информатика и информационные технологии": "ФИИТ",
        "Системный анализ и управление": "САУ"

    }
    not_kniit = {
        "Бизнес-информатика": "БИ",
        "Информационные системы и технологии": "ИСТ",
        "Прикладная математика и информатика": "ПМИ",
        "Математика и компьютерные науки": "МКН",
        "Прикладная информатика социологии": "ПРИ соц",
        "Прикладная информатика экономике": "ПРИ экон мех",
        "Прикладная информатика": "ПРИ мех",
        "Конструирование и технология электронных средств": "КТЭС",
        "Биотехнические системы и технологии": "БСТ",
        "Техносферная безопасность": "ТБ",
        "Геология": "Геол",
        "Механика и математическое моделирование": "МиММ",
        "Электроника и наноэлектроника": "ЭиН",
        "Экономика": "Экон",
        "Таможенное дело": "Тамож",
        "Менеджмент": "Менедж",
        "Медицинская биофизика": "МБ",
        "Медицинская кибернетика": "МК",
        "Прикладная геология": "ПГ",
        "Инфокоммуникационные технологии и системы связи": "ИТиСС",
        "Педагогическое образование Информатика": "ПО инф",
        "Педагогическое образование Математическое": "ПО мат",
        "Психолого-педагогическое образование": "ППО",
        "Социология": "Соц",
        "Сервис": "Серв"
    }

    for url in urls:
        req = requests.get(url)
        src = req.text
        soup = BeautifulSoup(src, 'lxml')
        tables = soup.find_all('table')
        # print(tables)
        for index, table in enumerate(tables):
            for row in table.find_all('tr'):
                cols = row.find_all('td')
                cols = [col.text.strip() for col in cols]
                if len(cols) > 9:
                    if cols[1] not in abituras:
                        cols[10] = ' '.join(cols[10].split())[4:]
                        cols[9] = int(cols[9])
                        i = 0
                        prior = [""] * 6
                        dop = []
                        tables = set()
                        s = cols[10].split()
                        # print(url)
                        ttl = soup.find_all(class_="svodka-table__info-wrap")[0]
                        ttl = ttl.find_all(class_="svodka-table__info-name")

                        ttl = (ttl[2].string, ttl[4].string, ttl[5].string.split(' ')[0])
                        while cols[9] >= len(prior):
                            prior.append('')

                        prior[cols[9] - 1] = kniit[ttl[0]]

                        if ttl[1] == "Заочная":
                            prior[cols[9] - 1] += " ЗО"
                            if ttl[2] != "Полное":
                                tables.add("zo")
                        if ttl[2] == "Бюджет/Общие" and ttl[1] != "Заочная":
                            tables.add("budj_obs")
                        if ttl[2] == "Бюджет/Особое":
                            tables.add("budj_osob")
                            prior[cols[9] - 1] += " ОП"
                        if ttl[2] == "Бюджет/Отдельная":
                            tables.add("budj_osob")
                            prior[cols[9] - 1] += " ОК"
                        if ttl[2] == "Полное":
                            tables.add("kom")
                            prior[cols[9] - 1] += " К"

                        if ttl[2] == "Целевой":
                            tables.add("cel")
                            prior[cols[9] - 1] += " Ц"
                        # print(cols[1])


                        while i < len(s):
                            if re.match(pattern, s[i]) is not None:
                                pr = int(s[i:][int(s[i:].index('Приоритет:') + 1)]) - 1
                                name = ""
                                j = i + 1
                                flag = True
                                while re.match(pattern2, s[j]) is None:
                                    if s[j] == "(Профиль:":
                                        flag = False
                                        if "Прикладная информатика" in name:
                                            name += " " + s[j + 4]
                                        if "Педагогическое образование" in name:
                                            name += " " + s[j + 1]

                                    if flag:
                                        name += " " + s[j]
                                    j += 1
                                name = name[1:]

                                if name in kniit:
                                    if prior[pr] != '':
                                        prior[pr]+=" "
                                    prior[pr] += kniit[name]
                                    if s[j] == "Бюджет/Общие":
                                        tables.add("budj_obs")
                                        if s[j + 2] == "(заоч.)":
                                            prior[pr] += " ЗО"
                                            tables.add("zo")

                                    elif s[j] == "Бюджет/Особое":
                                        tables.add("budj_osob")
                                        if s[j + 2] == "(заоч.)":
                                            prior[pr] += " ЗО"
                                            tables.add("zo")
                                        prior[pr] += " ОП"
                                    elif s[j] == "Бюджет/Отдельная":
                                        tables.add("budj_otdel")
                                        if s[j + 2] == "(заоч.)":
                                            prior[pr] += " ЗО"
                                            tables.add("zo")
                                        prior[pr] += " ОК"
                                    elif s[j] == "Полное":
                                        tables.add("kom")
                                        if s[j + 3] == "(заоч.)":
                                            prior[pr] += " ЗО"
                                        prior[pr] += " К"

                                    elif s[j] == "Целевой":
                                        tables.add("cel")
                                        if s[j + 2] == "(заоч.)":
                                            tables.add("zo")
                                            prior[pr] +=" ЗО"
                                        prior[pr] += " Ц"
                                else:
                                    if name in not_kniit:
                                        dop.append(str(pr + 1) + " - " + not_kniit[name])
                                    else:
                                        dop.append(name)
                                    if s[j] == "Бюджет/Общие":
                                        if s[j + 2] == "(заоч.)":
                                            dop.append(" ЗО")

                                    elif s[j] == "Бюджет/Особое":

                                        if s[j + 2] == "(заоч.)":
                                            dop.append(name + " ЗО")
                                        dop[-1] += " ОП"
                                    elif s[j] == "Бюджет/Отдельная":

                                        if s[j + 2] == "(заоч.)":
                                            dop.append(" ЗО")
                                        dop[-1] += " ОК"
                                    elif s[j] == "Полное":
                                        if s[j + 2] == "затрат":
                                            if s[j + 3] == "(заоч.)":
                                                dop.append(" ЗО")
                                        dop[-1] += " К"
                                    elif s[j] == "Целевой":
                                        if s[j + 2] == "(заоч.)":
                                            dop.append(" ЗО")
                                        dop[-1] += " Ц"
                                i = j
                            i += 1
                        abituras[remove_leading_zeros(cols[1])] = Abitura(cols[1:8], prior, tables, dop)

    for fl in abituras_old:
        wb = openpyxl.load_workbook("tables/" + fl + ".xlsx")
        sheet = wb.active
        for i in range(1, sheet.max_row + 1):
            row_data = []
            for j in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=i, column=j).value
                if cell_value is not None:
                    row_data.append(cell_value)
                else:
                    row_data.append('')
            abituras_old[fl][row_data[0]] = row_data

    for fl in spisok:
        g = 0

        workbook = xlsxwriter.Workbook("tables/" + fl + '.xlsx')
        worksheet = workbook.add_worksheet()
        for ab in abituras_old[fl]:
            if len(abituras_old[fl][ab]) < 20:
                abituras_old[fl][ab] = abituras_old[fl][ab] + [''] * (20 - len(abituras_old[fl][ab]))
            abitr = abituras_old[fl][ab]
            if ab not in abituras:
                for j in range(len(abitr)):
                    worksheet.write(g, j, abitr[j])
            else:
                row = abituras[ab].row
                for i in range(len(row)):
                    if row[i] == 'Нет данных' or row[i] == '---':
                        row[i] = ''
                    elif row[i] == 'Ориг.':
                        row[i] = 'Оригинал'
                    elif row[i] == 'Копия':
                        row[i] = 'Копия'
                if len(row) < 8:
                    row.insert(6, '')
                new_row = abituras_old[fl][ab][0:3] + abituras[ab].prior[0:5] + [
                    ' '.join(abituras[ab].prior[5:])] + row[1:] + [' '.join(abituras[ab].dop)] + abituras_old[fl][ ab][-3:]
                if abituras_old[fl][ab][15] != abituras[ab].row[7]:
                    print(ab, ": ",abituras_old[fl][ab][15], " -> ", abituras[ab].row[7])
                if abituras_old[fl][ab][13] != '':
                    if int(abituras_old[fl][ab][13]) < int(abituras[ab].row[5]):
                        print(int(abituras_old[fl][ab][13]) - int(int(abituras[ab].row[5])))
                if abituras[ab].row[7] == "Оригинал":
                    for num in range(5):
                        if abituras_old[fl][ab][3+num][:2] != abituras[ab].prior[num][:2]:
                            print(ab,abituras_old[fl][ab][3+num], "-> ",  abituras[ab].prior[num] )

                for j in range(len(new_row)):
                    worksheet.write(g, j, new_row[j])
            g += 1

        for ab in abituras:
            if fl in abituras[ab].tables and ab not in abituras_old[fl]:
                row = abituras[ab].row
                for i in range(len(row)):
                    if row[i] == 'Нет данных' or row[i] == '---':
                        row[i] = ''
                    elif row[i] == 'Ориг.':
                        row[i] = 'Оригинал'
                if len(row) < 8:
                    row.insert(6, '')
                new_row = row[0:1] + ['', ''] + abituras[ab].prior[0:5] + [
                    ' '.join(abituras[ab].prior[5:])] + row[1:] + [' '.join(abituras[ab].dop),
                                                                  '', 'госуслуги', 'Нет документа об образовании']
                for j in range(len(new_row)):
                    worksheet.write(g, j, new_row[j])
                g += 1
        workbook.close()



spisok = ["budj_obs", "budj_osob", "cel", "zo"]
func(spisok, urls)

urls = ["https://www.sgu.ru/svodka/1787", "https://www.sgu.ru/svodka/1788", "https://www.sgu.ru/svodka/1791",
        "https://www.sgu.ru/svodka/1792", "https://www.sgu.ru/svodka/1795", "https://www.sgu.ru/svodka/1794",
        "https://www.sgu.ru/svodka/1793"]
spisok = ["kom"]
func(spisok, urls)
